import openpyxl
import os
import logging
import pymysql

os.chdir('C:/tech/python/excelloc')
temp_comp=[]
std_cols=["Hospital_ID","Hospital_Name","Bed_count"]
rowcount=0
try:
    mydb = pymysql.connect(
        host="localhost",
        user="root",
        password="Bigboss@34",
        database="python_db"
    )
    mycursor = mydb.cursor()
    read_xl=openpyxl.load_workbook('NewHosp.xlsx')
    act_xl=read_xl.active
    for i in range(1,act_xl.max_column+1):
        temp_comp.append(act_xl.cell(row=1, column=i).value)
        #print(act_xl.cell(row=1, column=i).value)
    print(temp_comp)
    if temp_comp == std_cols:
        print("Good to go")
        for i in range(2,act_xl.max_row+1):
            print(act_xl.cell(row=i, column=1).value)

        for i in act_xl.iter_rows(min_row=2,values_only=True):
            print(i[0],i[1],i[2])
            mycursor.execute("INSERT INTO hospital values (%s, %s, %s)", (i[0],i[1],i[2]))
            rowcount=rowcount+1
        print("Total rows inserted {0}:".format(rowcount))
    else:
        logging.basicConfig(level=logging.INFO,filename='app.log', format='%(asctime)s :: %(levelname)s :: %(message)s')
        logging.info("Headers did not match for file {0}".format(read_xl))
except (Exception) as error:
    print(error)
    logging.error("Exception occurred", exc_info=True)
except (Exception, pymysql.err) as error:
       print("Error while getting data", error)
finally:
    mydb.commit()
    mycursor.close()
    mydb.close()